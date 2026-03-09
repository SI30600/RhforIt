import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

# Couleurs Somnum
BLEU_FONCE = "1B3A5F"  # Bleu marine foncé
TURQUOISE = "00A0B0"   # Turquoise/Cyan
BLANC = "FFFFFF"
GRIS_CLAIR = "F5F7FA"
GRIS_MOYEN = "E8ECF0"

def create_somnum_excel():
    wb = Workbook()
    
    # =============================================
    # FEUILLE 1: ARRIVÉE DE PERSONNEL
    # =============================================
    ws_arrivee = wb.active
    ws_arrivee.title = "ARRIVÉE"
    
    # Styles
    header_fill = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
    subheader_fill = PatternFill(start_color=TURQUOISE, end_color=TURQUOISE, fill_type="solid")
    input_fill = PatternFill(start_color=GRIS_CLAIR, end_color=GRIS_CLAIR, fill_type="solid")
    alt_row_fill = PatternFill(start_color=GRIS_MOYEN, end_color=GRIS_MOYEN, fill_type="solid")
    
    header_font = Font(bold=True, color=BLANC, size=14)
    subheader_font = Font(bold=True, color=BLANC, size=11)
    label_font = Font(bold=True, color=BLEU_FONCE, size=10)
    
    thin_border = Border(
        left=Side(style='thin', color=TURQUOISE),
        right=Side(style='thin', color=TURQUOISE),
        top=Side(style='thin', color=TURQUOISE),
        bottom=Side(style='thin', color=TURQUOISE)
    )
    
    # Largeur des colonnes
    ws_arrivee.column_dimensions['A'].width = 3
    ws_arrivee.column_dimensions['B'].width = 30
    ws_arrivee.column_dimensions['C'].width = 25
    ws_arrivee.column_dimensions['D'].width = 20
    ws_arrivee.column_dimensions['E'].width = 20
    
    # TITRE PRINCIPAL
    ws_arrivee.merge_cells('B2:E2')
    ws_arrivee['B2'] = "FORMULAIRE D'ARRIVÉE - SOMNUM"
    ws_arrivee['B2'].font = Font(bold=True, color=BLEU_FONCE, size=18)
    ws_arrivee['B2'].alignment = Alignment(horizontal='center')
    
    # ==================
    # SECTION 1: INFORMATIONS GÉNÉRALES
    # ==================
    row = 4
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "INFORMATIONS DU COLLABORATEUR"
    ws_arrivee[f'B{row}'].font = header_font
    ws_arrivee[f'B{row}'].fill = header_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    # Champs d'information
    fields = [
        ("Date de la demande", "JJ/MM/AAAA"),
        ("Date d'arrivée prévue", "JJ/MM/AAAA"),
        ("Nom", ""),
        ("Prénom", ""),
        ("Numéro de téléphone personnel", "(pour création compte Google)"),
        ("Centre de rattachement", "Sélectionner..."),
    ]
    
    row = 6
    for label, placeholder in fields:
        ws_arrivee[f'B{row}'] = label
        ws_arrivee[f'B{row}'].font = label_font
        ws_arrivee[f'B{row}'].alignment = Alignment(vertical='center')
        
        ws_arrivee.merge_cells(f'C{row}:E{row}')
        ws_arrivee[f'C{row}'] = placeholder if placeholder else ""
        ws_arrivee[f'C{row}'].fill = input_fill
        ws_arrivee[f'C{row}'].border = thin_border
        ws_arrivee[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws_arrivee.row_dimensions[row].height = 25
        row += 1
    
    # Liste déroulante pour Centre de rattachement
    centres = DataValidation(
        type="list",
        formula1='"Arles,Nîmes,Alès,Montpellier,Aubenas,Avignon,Le Mans,Lyon,Firminy,Rodez,Aurillac,Siège"',
        allow_blank=True
    )
    centres.error = "Veuillez sélectionner un centre dans la liste"
    centres.prompt = "Choisissez le centre de rattachement"
    ws_arrivee.add_data_validation(centres)
    centres.add(ws_arrivee['C11'])
    
    # ==================
    # SECTION 2: TYPE DE PERSONNEL
    # ==================
    row = 13
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "TYPE DE PERSONNEL (cocher une case)"
    ws_arrivee[f'B{row}'].font = header_font
    ws_arrivee[f'B{row}'].fill = header_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    row = 15
    ws_arrivee[f'B{row}'] = "Statut :"
    ws_arrivee[f'B{row}'].font = label_font
    
    # Liste déroulante pour Statut
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    statut_dv = DataValidation(
        type="list",
        formula1='"Salarié,Collaborateur"',
        allow_blank=True
    )
    ws_arrivee.add_data_validation(statut_dv)
    statut_dv.add(ws_arrivee[f'C{row}'])
    
    row = 16
    ws_arrivee[f'B{row}'] = "Fonction :"
    ws_arrivee[f'B{row}'].font = label_font
    
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    fonction_dv = DataValidation(
        type="list",
        formula1='"Médecin,IDE,Secrétaire,Technicien(ne) du sommeil,Assistant(e) médical(e),Sage-femme,Personnel Siège,Autre"',
        allow_blank=True
    )
    ws_arrivee.add_data_validation(fonction_dv)
    fonction_dv.add(ws_arrivee[f'C{row}'])
    
    # ==================
    # SECTION 3: BESOINS INFORMATIQUES
    # ==================
    row = 18
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "BESOINS INFORMATIQUES"
    ws_arrivee[f'B{row}'].font = header_font
    ws_arrivee[f'B{row}'].fill = header_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    # Sous-section: Ordinateur
    row = 20
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "ORDINATEUR"
    ws_arrivee[f'B{row}'].font = subheader_font
    ws_arrivee[f'B{row}'].fill = subheader_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    row = 22
    ws_arrivee[f'B{row}'] = "Besoin d'un ordinateur ?"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    oui_non_dv = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws_arrivee.add_data_validation(oui_non_dv)
    oui_non_dv.add(ws_arrivee[f'C{row}'])
    
    row = 23
    ws_arrivee[f'B{row}'] = "Type d'ordinateur :"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    type_pc_dv = DataValidation(type="list", formula1='"PC Fixe,PC Portable"', allow_blank=True)
    ws_arrivee.add_data_validation(type_pc_dv)
    type_pc_dv.add(ws_arrivee[f'C{row}'])
    
    row = 24
    ws_arrivee[f'B{row}'] = "Taille PC Portable :"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    taille_dv = DataValidation(type="list", formula1='"Petit (13-14"),Normal (15.6"),Grand (17")"', allow_blank=True)
    ws_arrivee.add_data_validation(taille_dv)
    taille_dv.add(ws_arrivee[f'C{row}'])
    
    row = 25
    ws_arrivee[f'B{row}'] = "Caméra intégrée ?"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    oui_non_dv2 = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws_arrivee.add_data_validation(oui_non_dv2)
    oui_non_dv2.add(ws_arrivee[f'C{row}'])
    
    # Sous-section: Téléphone
    row = 27
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "TÉLÉPHONE PORTABLE PROFESSIONNEL"
    ws_arrivee[f'B{row}'].font = subheader_font
    ws_arrivee[f'B{row}'].fill = subheader_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    row = 29
    ws_arrivee[f'B{row}'] = "Besoin d'un téléphone ?"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    oui_non_dv3 = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws_arrivee.add_data_validation(oui_non_dv3)
    oui_non_dv3.add(ws_arrivee[f'C{row}'])
    
    # Sous-section: Logiciels
    row = 31
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "LOGICIELS & ACCÈS"
    ws_arrivee[f'B{row}'].font = subheader_font
    ws_arrivee[f'B{row}'].fill = subheader_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    row = 33
    ws_arrivee[f'B{row}'] = "Logiciel métier :"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    logiciel_dv = DataValidation(type="list", formula1='"Nox,Lowenstein,Les deux,Aucun"', allow_blank=True)
    ws_arrivee.add_data_validation(logiciel_dv)
    logiciel_dv.add(ws_arrivee[f'C{row}'])
    
    row = 34
    ws_arrivee[f'B{row}'] = "Raccourci SomnoBook :"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    oui_non_dv4 = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws_arrivee.add_data_validation(oui_non_dv4)
    oui_non_dv4.add(ws_arrivee[f'C{row}'])
    
    row = 35
    ws_arrivee[f'B{row}'] = "Ajout d'imprimante :"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee[f'C{row}'] = "Sélectionner..."
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    oui_non_dv5 = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws_arrivee.add_data_validation(oui_non_dv5)
    oui_non_dv5.add(ws_arrivee[f'C{row}'])
    
    row = 37
    ws_arrivee[f'B{row}'] = "Accès aux sites Drive :"
    ws_arrivee[f'B{row}'].font = label_font
    ws_arrivee.merge_cells(f'C{row}:E{row}')
    ws_arrivee[f'C{row}'] = "(préciser les sites)"
    ws_arrivee[f'C{row}'].fill = input_fill
    ws_arrivee[f'C{row}'].border = thin_border
    ws_arrivee.row_dimensions[row].height = 30
    
    # Sous-section: Configuration automatique
    row = 39
    ws_arrivee.merge_cells(f'B{row}:E{row}')
    ws_arrivee[f'B{row}'] = "CONFIGURATION AUTOMATIQUE (rempli par IT)"
    ws_arrivee[f'B{row}'].font = subheader_font
    ws_arrivee[f'B{row}'].fill = subheader_fill
    ws_arrivee[f'B{row}'].alignment = Alignment(horizontal='center')
    
    auto_items = [
        "Création BAL Microsoft gratuite",
        "Intégration sur Intune",
        "Installation Antivirus",
        "Configuration Drive partagé",
        "Création signature mail",
        "Configuration Microsoft Entra ID",
        "Installation O365 + Chrome"
    ]
    
    row = 41
    for item in auto_items:
        ws_arrivee[f'B{row}'] = item
        ws_arrivee[f'B{row}'].font = Font(color=BLEU_FONCE, size=10)
        ws_arrivee[f'C{row}'] = "☐"
        ws_arrivee[f'C{row}'].font = Font(size=14)
        ws_arrivee[f'C{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # =============================================
    # FEUILLE 2: DÉPART DE PERSONNEL
    # =============================================
    ws_depart = wb.create_sheet("DÉPART")
    
    # Largeur des colonnes
    ws_depart.column_dimensions['A'].width = 3
    ws_depart.column_dimensions['B'].width = 30
    ws_depart.column_dimensions['C'].width = 25
    ws_depart.column_dimensions['D'].width = 20
    ws_depart.column_dimensions['E'].width = 20
    
    # TITRE PRINCIPAL
    ws_depart.merge_cells('B2:E2')
    ws_depart['B2'] = "FORMULAIRE DE DÉPART - SOMNUM"
    ws_depart['B2'].font = Font(bold=True, color=BLEU_FONCE, size=18)
    ws_depart['B2'].alignment = Alignment(horizontal='center')
    
    # Section: Informations du collaborateur
    row = 4
    ws_depart.merge_cells(f'B{row}:E{row}')
    ws_depart[f'B{row}'] = "INFORMATIONS DU COLLABORATEUR"
    ws_depart[f'B{row}'].font = header_font
    ws_depart[f'B{row}'].fill = header_fill
    ws_depart[f'B{row}'].alignment = Alignment(horizontal='center')
    
    depart_fields = [
        ("Date de la demande", "JJ/MM/AAAA"),
        ("Date de départ", "JJ/MM/AAAA"),
        ("Nom", ""),
        ("Prénom", ""),
        ("Centre de rattachement", "Sélectionner..."),
    ]
    
    row = 6
    for label, placeholder in depart_fields:
        ws_depart[f'B{row}'] = label
        ws_depart[f'B{row}'].font = label_font
        ws_depart[f'B{row}'].alignment = Alignment(vertical='center')
        
        ws_depart.merge_cells(f'C{row}:E{row}')
        ws_depart[f'C{row}'] = placeholder if placeholder else ""
        ws_depart[f'C{row}'].fill = input_fill
        ws_depart[f'C{row}'].border = thin_border
        ws_depart[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws_depart.row_dimensions[row].height = 25
        row += 1
    
    # Liste déroulante pour Centre
    centres_depart = DataValidation(
        type="list",
        formula1='"Arles,Nîmes,Alès,Montpellier,Aubenas,Avignon,Le Mans,Lyon,Firminy,Rodez,Aurillac,Siège"',
        allow_blank=True
    )
    ws_depart.add_data_validation(centres_depart)
    centres_depart.add(ws_depart['C10'])
    
    # Section: Actions de départ
    row = 12
    ws_depart.merge_cells(f'B{row}:E{row}')
    ws_depart[f'B{row}'] = "ACTIONS À RÉALISER PAR L'IT"
    ws_depart[f'B{row}'].font = header_font
    ws_depart[f'B{row}'].fill = header_fill
    ws_depart[f'B{row}'].alignment = Alignment(horizontal='center')
    
    depart_actions = [
        ("Sauvegarde des mails", "Dans: ONEDRIVE KENNEDY\\SIEGE\\IT"),
        ("Suppression compte Microsoft", ""),
        ("Libération licence Microsoft", ""),
        ("Récupération matériel", "PC / Téléphone"),
        ("Désactivation accès Drive", ""),
        ("Suppression Intune", ""),
    ]
    
    row = 14
    for action, note in depart_actions:
        ws_depart[f'B{row}'] = action
        ws_depart[f'B{row}'].font = label_font
        
        ws_depart[f'C{row}'] = "☐ Fait"
        ws_depart[f'C{row}'].font = Font(size=11)
        ws_depart[f'C{row}'].fill = input_fill
        ws_depart[f'C{row}'].border = thin_border
        
        if note:
            ws_depart[f'D{row}'] = note
            ws_depart[f'D{row}'].font = Font(italic=True, color="666666", size=9)
        
        ws_depart.row_dimensions[row].height = 25
        row += 1
    
    # Commentaires
    row += 2
    ws_depart[f'B{row}'] = "Commentaires :"
    ws_depart[f'B{row}'].font = label_font
    
    row += 1
    ws_depart.merge_cells(f'B{row}:E{row+2}')
    ws_depart[f'B{row}'].fill = input_fill
    ws_depart[f'B{row}'].border = thin_border
    
    # =============================================
    # FEUILLE 3: LÉGENDE / AIDE
    # =============================================
    ws_aide = wb.create_sheet("AIDE")
    
    ws_aide.column_dimensions['A'].width = 3
    ws_aide.column_dimensions['B'].width = 50
    
    ws_aide.merge_cells('B2:D2')
    ws_aide['B2'] = "GUIDE D'UTILISATION"
    ws_aide['B2'].font = Font(bold=True, color=BLEU_FONCE, size=16)
    
    aide_content = [
        "",
        "COMMENT REMPLIR CE FORMULAIRE :",
        "",
        "1. Les champs gris clair sont à remplir",
        "2. Cliquez sur les cellules avec 'Sélectionner...' pour voir les options",
        "3. Les cases ☐ sont à cocher manuellement (remplacer par ☑)",
        "",
        "CENTRES SOMNUM :",
        "• Arles, Nîmes, Alès, Montpellier",
        "• Aubenas, Avignon, Le Mans, Lyon",
        "• Firminy, Rodez, Aurillac, Siège",
        "",
        "LOGICIELS MÉTIER :",
        "• Nox : logiciel de polysomnographie",
        "• Lowenstein : logiciel de ventilation",
        "",
        "PRÉPARATION POSTE STANDARD :",
        "• O365 (Office 365)",
        "• Microsoft Entra ID (authentification)",
        "• Chrome (navigateur)",
        "• Antivirus",
        "• Drive partagé",
    ]
    
    row = 4
    for line in aide_content:
        ws_aide[f'B{row}'] = line
        if line.startswith("COMMENT") or line.startswith("CENTRES") or line.startswith("LOGICIELS") or line.startswith("PRÉPARATION"):
            ws_aide[f'B{row}'].font = Font(bold=True, color=TURQUOISE, size=11)
        else:
            ws_aide[f'B{row}'].font = Font(color=BLEU_FONCE, size=10)
        row += 1
    
    # Sauvegarder
    output_path = "/app/formulaire_rh_somnum.xlsx"
    wb.save(output_path)
    print(f"Fichier Excel créé: {output_path}")
    return output_path

if __name__ == "__main__":
    create_somnum_excel()
